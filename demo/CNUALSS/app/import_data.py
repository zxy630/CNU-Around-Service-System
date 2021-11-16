import os

from flask_script import Command
import openpyxl

from app.config import BASE_DIR
from app.models import Cate
from app.models import Beauty
from app.models import Recreation
from app import db


class ImportData(Command):
    def run(self):
        print('导入数据')
        dir = os.path.join(BASE_DIR, 'excel')
        for file_name in os.listdir(dir):
            file_path = os.path.join(dir, file_name)
            self.save_to_mysql(file_path)
        # file_path = os.path.join(dir, 'Beauty')  # 连接路径和文件名
        # self.save_to_mysql(file_path)
        # file_path = os.path.join(dir, 'Recreation')  # 连接路径和文件名
        # self.save_to_mysql(file_path)
        print('导入完成')

    def save_to_mysql(self, file_path):
        fileds = ['plate', 'ID', 'name', 'label', 'opening_hours', 'person_price', 'telephone',
                  'address', 'special', 'img_url']
        # 导入excel
        wb = openpyxl.load_workbook(file_path)
        # 选择worksheet
        index = wb.sheetnames
        ws = wb[index[0]]  # 第一个sheet
        # 迭代行
        for row in ws.iter_rows(min_row=2):
            data = [cell.value for cell in row]
            dict_val = dict(zip(fileds, data))
            # 存入美食Cate表
            if dict_val.get('plate') == 1:
                cate = Cate(**dict_val)
                self.save_cate(cate)
            elif dict_val.get('plate') == 2:
                beauty = Beauty(**dict_val)
                self.save_beauty(beauty)
            elif dict_val.get('plate') == 3:
                recreation = Recreation(**dict_val)
                self.save_cate(recreation)
        wb.close()

    def save_cate(self, cate):
        try:
            db.session.merge(cate)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(e)
            return e

    def save_beauty(self, beauty):
        try:
            db.session.merge(beauty)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(e)
            return e

    def save_recreation(self, recreation):
        try:
            db.session.merge(recreation)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(e)
            return e


if __name__ == '__main__':
    print(BASE_DIR)
